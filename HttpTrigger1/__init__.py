import logging
import os
import azure.functions as func
import requests
import json
import io
from urllib.parse import quote
from docx import Document
from openpyxl import load_workbook
from PyPDF2 import PdfReader

SHAREPOINT_SITE_ID = "lctoyotaoutlook.sharepoint.com,2c8b9562-cf1a-40c2-8a1a-31c6d62b59d6,8b64fd9c-db3b-480e-b34a-20d0bb183edd"

CONTENT_TYPE_MAP = {
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
    "application/pdf": ".pdf",
    "text/plain": ".txt"
}

def parse_docx(content_bytes):
    return "\n".join([p.text for p in Document(io.BytesIO(content_bytes)).paragraphs])

def parse_xlsx(content_bytes):
    wb = load_workbook(io.BytesIO(content_bytes), data_only=True)
    content = ""
    for sheet in wb.worksheets:
        content += f"Sheet: {sheet.title}\n"
        for row in sheet.iter_rows(values_only=True):
            content += "\t".join(str(cell or "") for cell in row) + "\n"
    return content

def parse_pdf(content_bytes):
    reader = PdfReader(io.BytesIO(content_bytes))
    return "".join([page.extract_text() or "" for page in reader.pages])

def parse_text(content_bytes):
    return content_bytes.decode("utf-8", errors="replace")

def detect_extension(content_type):
    return CONTENT_TYPE_MAP.get(content_type, ".txt")

def create_json_response(data, status=200):
    return func.HttpResponse(
        json.dumps(data),
        status_code=status,
        mimetype="application/json",
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET",
            "Access-Control-Allow-Headers": "Content-Type"
        }
    )

def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("SharePoint handler triggered.")

    file_id = req.params.get("fileId")
    summary = req.params.get("summary") == "true"
    debug = req.params.get("debug") == "true"
    list_files = req.params.get("list") == "true"

    tenant_id = os.getenv("TENANT_ID")
    client_id = os.getenv("CLIENT_ID")
    client_secret = os.getenv("CLIENT_SECRET")

    if not tenant_id or not client_id or not client_secret:
        return create_json_response({"error": "Missing environment variables."}, 500)

    # Get access token
    token_resp = requests.post(
        f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token",
        data={
            "client_id": client_id,
            "client_secret": client_secret,
            "scope": "https://graph.microsoft.com/.default",
            "grant_type": "client_credentials"
        }
    )

    if token_resp.status_code != 200:
        return create_json_response({"error": "Failed to get access token.", "details": token_resp.text}, 401)

    token_json = token_resp.json()
    token = token_json.get("access_token")

    if not token:
        return create_json_response({"error": "Token missing from response.", "details": token_json}, 401)

    headers = {"Authorization": f"Bearer {token}"}

    # === Case 1: List files ===
    if list_files:
        list_url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_ID}/drive/root/children?$top=100"
        resp = requests.get(list_url, headers=headers)
        if resp.status_code != 200:
            logging.error(f"List error: {resp.text}")
            return create_json_response({"error": "Failed to list files.", "details": resp.text}, resp.status_code)

        files = []
        for item in resp.json().get("value", []):
            if "file" not in item:
                continue
            path = item["parentReference"]["path"].split("/root:/")[-1] + "/" + item["name"]
            files.append({
                "name": item["name"],
                "fileId": item["id"],
                "type": item["file"]["mimeType"].split("/")[-1],
                "lastModified": item.get("lastModifiedDateTime", ""),
                "path": path
            })

        return create_json_response(files)

    # === Case 2: Fetch file by fileId ===
    elif file_id:
        content_url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_ID}/drive/items/{file_id}/content"
        logging.info(f"Fetching file content from: {content_url}")

        file_resp = requests.get(content_url, headers=headers)
        if file_resp.status_code != 200:
            logging.error("File fetch failed.")
            logging.error(f"URL: {content_url}")
            logging.error(f"fileId: {file_id}")
            logging.error(f"Response code: {file_resp.status_code}")
            logging.error(f"Response text: {file_resp.text}")

            return create_json_response({
                "error": "Failed to fetch file content.",
                "reason": "Microsoft Graph returned an error.",
                "fileId": file_id,
                "graphUrl": content_url,
                "statusCode": file_resp.status_code,
                "graphResponse": file_resp.text
            }, file_resp.status_code)

        content_bytes = file_resp.content
        content_type = file_resp.headers.get("Content-Type", "")
        ext = detect_extension(content_type)

        try:
            if ext == ".docx":
                text = parse_docx(content_bytes)
            elif ext == ".xlsx":
                text = parse_xlsx(content_bytes)
            elif ext == ".pdf":
                text = parse_pdf(content_bytes)
            else:
                text = parse_text(content_bytes)
        except Exception as e:
            return create_json_response({"error": "Failed to parse file.", "details": str(e)}, 500)

        if summary:
            text = text[:2000] + "\n...\n[Content truncated]"

        return create_json_response({
            "content": text,
            "contentUrl": content_url if debug else None
        })

    # === Case 3: No valid query ===
    else:
        return create_json_response({
            "status": "ready",
            "message": "Use 'list=true' to browse files, and 'fileId=...' to retrieve content. 'filename' is no longer supported."
        })
